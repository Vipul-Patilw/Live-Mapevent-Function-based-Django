from mapeventApp.models import Event,Login
from django.shortcuts import redirect, render
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
def booking_admin_view(request):
	if request.method == "POST":
		event = request.POST.get('event')
		if event:
			events1 = Event.objects.filter(event=event).all()
		else:
			events1 = Event.objects.all()
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = f"{event} Event Data"
		headers = [ 'Event Name', 'Event Address', 'Event Date & Time', 'User Name','Booking Date', 'Booking Time']
		for col_num, column_title in enumerate(headers, 1):
			col_letter = get_column_letter(col_num)
			ws[f'{col_letter}1'] = column_title
		for row_num, event_obj in enumerate(events1, 2):
			ws[f'A{row_num}'] = str(event_obj.event)
			
			ws[f'B{row_num}'] = event_obj.eventaddress
			ws[f'C{row_num}'] = str(event_obj.dtime)
			ws[f'D{row_num}'] = event_obj.full_name
			ws[f'E{row_num}'] = str(event_obj.date)
			ws[f'F{row_num}'] = str(event_obj.time)
		response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
		response['Content-Disposition'] = f'attachment; filename={event}_event_data.xlsx'
		wb.save(response)
		return response
	eventscount = Event.objects.all()
	events2 = Event.objects.all().values_list('event', flat=True).distinct()
#	events2 = Event.objects.values('event' ).annotate(Count('id')).order_by().filter(id__count__gt=1)
	try:
		events3 = Event.objects.filter(event=events2[0]).count()
		
	except:
		return render(request,'booking.html')

	
	main = {'events1':events2,'events3':events3,'eventcount':eventscount}
	
	return render (request,'booking.html',main)

def booking_user_view(request, email):
		events = Event.objects.filter(email=email).all()
		
		
		return render (request,'booking.html', {'events':events,})
	